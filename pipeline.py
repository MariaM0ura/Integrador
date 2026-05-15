"""
pipeline.py
===========
Orquestrador principal do SellersFlow.

Une Reader → Mapper → RuleFiller → AI Fill → Filler em um pipeline único.
É o único ponto de entrada que o app.py (Streamlit) precisa chamar.

Design:
  - Stateless: cada chamada a run() é independente
  - Retorna PipelineResult com todos os artefatos e logs
  - Suporta modo "dry_run" (mapeia mas não grava arquivo)
  - Suporta enriquecimento por IA (opt-in)
  - Suporta análise de instruções do template (opt-in)
  - Suporta qualquer marketplace como ORIGEM (não apenas Amazon)

Fases (com use_instructions=True):
  FASE 1: Mapeamento de colunas (estratégias: aprendido → fixo → similaridade → IA)
  FASE 2: RuleBasedFiller (lookup de valores aceitos, concatenação, herança de exemplo)
  FASE 3: AI fill para obrigatórias ainda vazias
  FASE 4: Herança de exemplos para opcionais ainda vazias
"""

from __future__ import annotations

import io
import logging
import time
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd

from core.reader import AmazonSheetReader, AmazonReadResult
from core.source_reader import MarketplaceSourceReader, SourceReadResult
from core.mapper import ColumnMapper, MappingResult, FieldMappingDecision, _normalize as _mapper_normalize
from core.filler import (
    MarketplaceFiller,
    FillResult,
    guess_best_header_row,
    guess_best_header_row_from_sheet_xml,
    _find_sheet_zip_path,
)
from core.enricher import EnricherPipeline
from core.enricher.enricher_pipeline import detect_empty_fields
from core.instruction_parser import InstructionParser, ColumnRule
from core.rule_filler import RuleBasedFiller
from ai.ai_engine import AIEngine

logger = logging.getLogger(__name__)

# ─── Caminho do banco de aprendizado ─────────────────────────────────────────

DEFAULT_DB_PATH = Path(__file__).parent / "data" / "mappings_db" / "learned.json"


# ─── Dataclass de resultado ───────────────────────────────────────────────────

@dataclass
class PipelineResult:
    # Metadados
    marketplace: str
    elapsed_seconds: float

    # Resultados por etapa
    read_result: Optional[AmazonReadResult] = None
    mapping_result: Optional[MappingResult] = None
    fill_result: Optional[FillResult] = None

    # Cobertura por fase (preenchida quando use_instructions=True)
    # {"fase1_mapping": 0.65, "fase2_rule": 0.12, "fase3_ai": 0.05, "total": 0.82}
    phase_coverage: dict[str, float] = field(default_factory=dict)

    # Flags de alto nível
    success: bool = False
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def output_path(self) -> Optional[str]:
        return self.fill_result.output_path if self.fill_result else None

    @property
    def amazon_df(self) -> Optional[pd.DataFrame]:
        return self.read_result.df if self.read_result else None

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)


# ─── Pipeline ─────────────────────────────────────────────────────────────────

class SellersFlowPipeline:
    """
    Pipeline principal do SellersFlow.

    Uso típico:
        pipeline = SellersFlowPipeline()
        result = pipeline.run(
            amazon_file=bytes_io_origem,
            template_file=bytes_io_template,
            marketplace="Shopee",           # marketplace DESTINO
            source_marketplace="Amazon",    # marketplace ORIGEM (default)
            use_ai=True,
        )
        if result.success:
            print(result.output_path)
    """

    def __init__(
        self,
        db_path: Optional[Path] = None,
        output_dir: Optional[str] = None,
    ):
        self._reader = AmazonSheetReader()
        self._source_reader = MarketplaceSourceReader()
        self._mapper = ColumnMapper(db_path=db_path or DEFAULT_DB_PATH)
        self._filler = MarketplaceFiller()
        self._ai = AIEngine()
        self._instruction_parser = InstructionParser()
        self._rule_filler = RuleBasedFiller()
        self._output_dir = output_dir
        self._enricher_pipeline: Optional[EnricherPipeline] = None

    # ── Pública ───────────────────────────────────────────────────────────────

    def run(
        self,
        amazon_file,
        template_file,
        marketplace: str,
        use_ai: bool = False,
        enrich_ai: bool = False,
        dry_run: bool = False,
        source_marketplace: str = "Amazon",
        use_instructions: bool = False,
    ) -> PipelineResult:
        """
        Executa o pipeline completo.

        Args:
            amazon_file:        BytesIO da planilha de ORIGEM.
            template_file:      BytesIO do template do marketplace DESTINO.
            marketplace:        Marketplace DESTINO ("Shopee", "Temu", "Amazon"...).
            use_ai:             Se True, usa IA como fallback de mapeamento.
            enrich_ai:          Se True, aplica enriquecimento de conteúdo via IA.
            dry_run:            Se True, não grava arquivo de saída.
            source_marketplace: Marketplace ORIGEM dos dados (default "Amazon").
            use_instructions:   Se True, ativa as fases 2-4 (RuleFiller + AI fill
                                + herança de exemplos) usando as abas de instrução
                                do template.

        Returns:
            PipelineResult completo (com phase_coverage preenchido se use_instructions).
        """
        t0 = time.perf_counter()
        result = PipelineResult(marketplace=marketplace, elapsed_seconds=0.0)

        # ── Etapa 1: Leitura da planilha de origem ────────────────────────
        logger.info("[Pipeline] Lendo origem '%s'...", source_marketplace)

        if source_marketplace == "Amazon":
            read_result = self._reader.read(amazon_file)
            result.read_result = read_result
            if read_result.has_errors:
                result.errors.extend(read_result.errors)
                result.elapsed_seconds = time.perf_counter() - t0
                return result
            result.warnings.extend(read_result.warnings)
            amazon_df = read_result.df
        else:
            src_result = self._source_reader.read(amazon_file, source_marketplace)
            # Adapta SourceReadResult para o campo read_result esperado
            result.read_result = AmazonReadResult(
                df=src_result.df,
                language="BR",
                sheet_name=src_result.sheet_name,
                total_rows=src_result.total_rows,
                valid_rows=src_result.valid_rows,
                warnings=src_result.warnings,
                errors=src_result.errors,
            )
            if src_result.has_errors:
                result.errors.extend(src_result.errors)
                result.elapsed_seconds = time.perf_counter() - t0
                return result
            result.warnings.extend(src_result.warnings)
            # Normaliza colunas para chaves semânticas
            amazon_df = self._mapper.normalize_source_df(src_result.df, source_marketplace)
            # Atualiza read_result.df para que source_col bata com df.columns
            # no preview e no filler
            result.read_result.df = amazon_df
            logger.info(
                "[Pipeline] Origem '%s' normalizada: %d colunas semânticas",
                source_marketplace, len(amazon_df.columns),
            )

        # ── Etapa 2: Enriquecimento por IA (opcional) ─────────────────────
        if enrich_ai:
            logger.info("[Pipeline] Enriquecendo com IA (legado)...")
            amazon_df = self._apply_enrichment(amazon_df, marketplace)

        # ── Etapa 3: Obter cabeçalhos do template ─────────────────────────
        _template_ext = None
        if hasattr(template_file, "name") and template_file.name:
            _template_ext = Path(template_file.name).suffix.lower() or None
        if hasattr(template_file, "seek"):
            template_file.seek(0)
            _template_bytes = template_file.read()
        else:
            _template_bytes = Path(template_file).read_bytes()

        logger.info("[Pipeline] Lendo template %s...", marketplace)
        dest_headers = self._read_template_headers(_template_bytes, marketplace)
        if dest_headers is None:
            result.errors.append("Não foi possível ler os cabeçalhos do template.")
            result.elapsed_seconds = time.perf_counter() - t0
            return result

        # ── Etapa 4: Mapeamento (FASE 1) ─────────────────────────────────
        logger.info("[Pipeline] FASE 1 — Construindo mapeamento...")
        ai_engine = self._ai if use_ai else None
        mapping = self._mapper.build_mapping(
            amazon_df=amazon_df,
            dest_headers=dest_headers,
            marketplace=marketplace,
            ai_engine=ai_engine,
        )
        result.mapping_result = mapping

        # ── Fases 2-4: Análise de instruções (opt-in) ─────────────────
        if use_instructions:
            amazon_df, mapping = self._apply_instruction_phases(
                amazon_df=amazon_df,
                mapping=mapping,
                dest_headers=dest_headers,
                template_bytes=_template_bytes,
                marketplace=marketplace,
                use_ai=use_ai,
                result=result,
            )
            result.mapping_result = mapping

        if dry_run:
            result.success = True
            result.elapsed_seconds = time.perf_counter() - t0
            return result

        # ── Etapa 5: Enriquecimento em cascata (Defaults → Regex → LLM) ──────
        import os
        llm_api_key = os.getenv("ANTHROPIC_API_KEY")
        enricher = EnricherPipeline(
            use_llm=use_ai and bool(llm_api_key),
            llm_api_key=llm_api_key,
        )
        dest_header_names = list(dest_headers.values())
        col_mapping = {dest: src for dest, src in mapping.column_map.items()} if hasattr(mapping, "column_map") else {}

        enriched_rows = []
        for _, row in amazon_df.iterrows():
            product = row.to_dict()
            empty = detect_empty_fields(product, dest_header_names, col_mapping)
            enrichment = enricher.enrich_product(
                product=product,
                empty_fields=empty,
                destination_marketplace=marketplace.lower(),
            )
            enriched_rows.append(enrichment.enriched_product)
            if enrichment.still_empty:
                title = product.get("title") or product.get("titulo") or product.get("item name", "?")
                result.warnings.append(
                    f"'{title}': campos não preenchidos: {enrichment.still_empty}"
                )

        if enriched_rows:
            amazon_df = pd.DataFrame(enriched_rows)
            logger.info("[Pipeline] Enriquecimento concluído para %d produtos.", len(enriched_rows))

        # Mantém read_result.df alinhado ao DataFrame usado no fill/preview
        # (colunas virtuais __rf__/__ai__ das fases 2–4 e enriquecimento).
        if result.read_result is not None:
            result.read_result.df = amazon_df

        # ── Etapa 6: Preenchimento ────────────────────────────────────────
        logger.info("[Pipeline] Preenchendo template...")
        fill_result = self._filler.fill(
            amazon_df=amazon_df,
            mapping=mapping,
            template_file=io.BytesIO(_template_bytes),
            output_dir=self._output_dir,
            template_ext=_template_ext,
        )
        result.fill_result = fill_result

        if fill_result.has_errors:
            result.errors.extend(fill_result.errors)
        else:
            result.success = True

        result.warnings.extend(fill_result.warnings)
        result.elapsed_seconds = round(time.perf_counter() - t0, 2)
        logger.info("[Pipeline] Concluído em %.2fs", result.elapsed_seconds)
        return result

    def learn_mapping(
        self, marketplace: str, dest_col: str, source_col: str
    ) -> None:
        """Persiste uma decisão de mapeamento confirmada pelo usuário."""
        self._mapper.learn(marketplace, dest_col, source_col)

    # ── Fases 2-4: instrução + regra + IA ────────────────────────────────────

    def _apply_instruction_phases(
        self,
        amazon_df: pd.DataFrame,
        mapping: MappingResult,
        dest_headers: dict[int, str],
        template_bytes: bytes,
        marketplace: str,
        use_ai: bool,
        result: PipelineResult,
    ) -> tuple[pd.DataFrame, MappingResult]:
        """
        FASE 2: RuleBasedFiller — lookup, concatenação, herança de exemplo
        FASE 3: AI fill — para obrigatórias ainda vazias (se use_ai=True)
        FASE 4: Herança de exemplos — para opcionais ainda vazias

        dest_headers: {excel_col_idx: col_name} — chave é o índice real no Excel.
        Retorna (amazon_df aumentado, mapping atualizado).
        """
        from core.filler import MARKETPLACE_CONFIG

        config = MARKETPLACE_CONFIG.get(marketplace, {})
        header_row = int(config.get("header_row", 3))

        # Aba de dados para o parser
        try:
            data_sheet = self._filler._resolve_sheet_name(template_bytes, config, marketplace)
        except Exception:
            data_sheet = ""
        if not data_sheet:
            logger.warning("[Fases 2-4] Aba não determinada para '%s'", marketplace)
            return amazon_df, mapping

        # ── Parse de instruções ───────────────────────────────────────────────
        logger.info("[Pipeline] FASE 2-4 — Parseando instruções de '%s'...", marketplace)
        col_rules = self._instruction_parser.parse(
            template_bytes, marketplace, data_sheet, header_row
        )
        example_rows = self._collect_example_rows(template_bytes, marketplace)

        # ── Constrói mapa bidirecional excel_col_idx ↔ col_name ──────────────
        # dest_headers: {excel_col_idx: col_name}
        # Decisões são geradas em ordem de dest_headers.items()
        col_name_to_excel_idx: dict[str, int] = {v: k for k, v in dest_headers.items()}

        # Colunas sem source_idx = candidatas para preenchimento por regra
        # Usamos excel_col_idx como chave (o que o index_map usa)
        unmapped: dict[int, str] = {
            col_name_to_excel_idx[d.dest_col]: d.dest_col
            for d in mapping.decisions
            if d.source_idx is None and d.dest_col in col_name_to_excel_idx
        }

        total_dest = len(mapping.decisions)
        mapped_before = total_dest - len(unmapped)
        result.phase_coverage["fase1_mapping"] = round(mapped_before / total_dest, 3) if total_dest else 0.0
        logger.info(
            "[Pipeline] FASE 1 cobertura: %d/%d (%.0f%%)",
            mapped_before, total_dest, result.phase_coverage["fase1_mapping"] * 100,
        )

        if not unmapped:
            for k in ("fase2_rule", "fase3_ai", "fase4_exemplo"):
                result.phase_coverage[k] = 0.0
            result.phase_coverage["total"] = result.phase_coverage["fase1_mapping"]
            return amazon_df, mapping

        # Índice reverso: dec_list_pos de cada dest_col
        dec_idx_by_col: dict[str, int] = {
            d.dest_col: i for i, d in enumerate(mapping.decisions)
        }

        def _update_mapping(
            aug_df: pd.DataFrame,
            col_name: str,
            col_key: str,
            strategy: str,
            confidence: float,
            notes: str,
        ) -> bool:
            """Atualiza decision + index_map com a nova coluna virtual. Retorna True se ok."""
            excel_idx = col_name_to_excel_idx.get(col_name)
            dec_i = dec_idx_by_col.get(col_name)
            if excel_idx is None or dec_i is None or col_key not in aug_df.columns:
                return False
            src_idx = aug_df.columns.get_loc(col_key)
            mapping.decisions[dec_i] = FieldMappingDecision(
                dest_col=col_name,
                source_col=col_key,
                source_idx=src_idx,
                strategy=strategy,
                confidence=confidence,
                notes=notes,
            )
            mapping.index_map[excel_idx] = src_idx
            if col_name in mapping.unmapped_dest:
                mapping.unmapped_dest.remove(col_name)
            return True

        # ── FASE 2: RuleBasedFiller ───────────────────────────────────────────
        logger.info("[Pipeline] FASE 2 — RuleBasedFiller (%d cols)...", len(unmapped))
        aug_df, rf_idx_map = self._rule_filler.build_augmented_df(
            amazon_df=amazon_df,
            unmapped_dest=unmapped,          # {excel_col_idx: col_name}
            col_rules=col_rules,
            example_rows=example_rows,
        )

        fase2_new = 0
        for excel_col_idx, col_name in unmapped.items():
            if excel_col_idx in rf_idx_map:
                col_key = f"__rf__{col_name}"
                if _update_mapping(aug_df, col_name, col_key, "rule", 0.75,
                                   "Preenchido por RuleBasedFiller (instrução do template)."):
                    fase2_new += 1

        result.phase_coverage["fase2_rule"] = round(fase2_new / total_dest, 3) if total_dest else 0.0
        logger.info("[Pipeline] FASE 2: +%d colunas (%.0f%%)", fase2_new, result.phase_coverage["fase2_rule"] * 100)

        # ── FASE 3: AI fill para obrigatórias ainda vazias ───────────────────
        fase3_new = 0
        if use_ai:
            logger.info("[Pipeline] FASE 3 — AI fill obrigatórias...")
            still_unmapped_3 = {
                ei: col
                for ei, col in unmapped.items()
                if ei not in rf_idx_map
                and (rule := col_rules.get(col)) and rule.obrigatorio
            }
            if still_unmapped_3:
                ai_vals: dict[str, list] = {col: [] for col in still_unmapped_3.values()}
                for _, row in aug_df.iterrows():
                    src = {k: v for k, v in row.to_dict().items() if not str(k).startswith("__")}
                    for col_name in still_unmapped_3.values():
                        rule = col_rules.get(col_name)
                        if not rule:
                            ai_vals[col_name].append(None)
                            continue
                        ai_res = self._ai.analyze_instruction_and_fill(
                            column_name=col_name,
                            instruction_text=rule.regra,
                            accepted_values=rule.valores_aceitos,
                            examples=[rule.exemplo] if rule.exemplo else [],
                            amazon_row_data=src,
                            marketplace=marketplace,
                        )
                        ai_vals[col_name].append(ai_res.get("value") if ai_res else None)

                for col_name in still_unmapped_3.values():
                    vals = ai_vals[col_name]
                    if any(v for v in vals):
                        col_key = f"__ai__{col_name}"
                        aug_df[col_key] = vals
                        if _update_mapping(aug_df, col_name, col_key, "ai_instruction", 0.7,
                                           "Preenchido por IA usando instrução do template."):
                            fase3_new += 1

        result.phase_coverage["fase3_ai"] = round(fase3_new / total_dest, 3) if total_dest else 0.0
        logger.info("[Pipeline] FASE 3: +%d colunas (%.0f%%)", fase3_new, result.phase_coverage["fase3_ai"] * 100)

        # ── FASE 4: Herança de exemplos para opcionais ───────────────────────
        fase4_new = 0
        if example_rows:
            from core.rule_filler import _from_example
            filled_so_far = set(rf_idx_map.keys())
            still_unmapped_4 = {
                ei: col
                for ei, col in unmapped.items()
                if ei not in filled_so_far
                and (di := dec_idx_by_col.get(col)) is not None
                and mapping.decisions[di].source_idx is None
            }
            for col_name in still_unmapped_4.values():
                ex_val = _from_example(col_name, example_rows)
                if ex_val:
                    col_key = f"__ex__{col_name}"
                    aug_df[col_key] = [ex_val] * len(aug_df)
                    if _update_mapping(aug_df, col_name, col_key, "exemplo", 0.5,
                                       "Herdado da aba de exemplos do template."):
                        fase4_new += 1

        result.phase_coverage["fase4_exemplo"] = round(fase4_new / total_dest, 3) if total_dest else 0.0

        total_mapped = mapped_before + fase2_new + fase3_new + fase4_new
        result.phase_coverage["total"] = round(total_mapped / total_dest, 3) if total_dest else 0.0

        # Mandatory coverage: uses REQUIRED_FIELDS for the destination marketplace
        from core.mapper import REQUIRED_FIELDS
        required_list = REQUIRED_FIELDS.get(marketplace, [])
        if required_list:
            dec_by_name = {d.dest_col: d for d in mapping.decisions}
            mandatory_mapped = sum(
                1 for col in required_list
                if dec_by_name.get(col) and dec_by_name[col].source_idx is not None
            )
            result.phase_coverage["mandatory_coverage"] = round(
                mandatory_mapped / len(required_list), 3
            )
        else:
            result.phase_coverage["mandatory_coverage"] = result.phase_coverage["total"]

        logger.info(
            "[Pipeline] Cobertura final: %d/%d (%.0f%%) — F1=%d F2=%d F3=%d F4=%d | mandatory=%.0f%%",
            total_mapped, total_dest, result.phase_coverage["total"] * 100,
            mapped_before, fase2_new, fase3_new, fase4_new,
            result.phase_coverage["mandatory_coverage"] * 100,
        )
        return aug_df, mapping

    def _collect_example_rows(
        self, template_bytes: bytes, marketplace: str
    ) -> list[dict]:
        """Extrai linhas de exemplo das abas de exemplo do template."""
        from core.instruction_parser import _load_wb, _EXAMPLE_PATTERNS, _norm
        import warnings as _w
        _w.filterwarnings("ignore")

        rows: list[dict] = []
        try:
            wb = _load_wb(template_bytes)
            for sheet_name in wb.sheetnames:
                sn = _norm(sheet_name)
                if not any(p in sn for p in _EXAMPLE_PATTERNS):
                    continue
                ws = wb[sheet_name]
                all_rows = list(ws.iter_rows(values_only=True))
                if len(all_rows) < 2:
                    continue
                # Find header row
                hdr_idx = 0
                best = 0
                for i, r in enumerate(all_rows[:5]):
                    cnt = sum(1 for v in r if v)
                    if cnt > best:
                        best, hdr_idx = cnt, i
                headers = [str(v).strip() if v else "" for v in all_rows[hdr_idx]]
                for row in all_rows[hdr_idx + 1: hdr_idx + 8]:
                    if not row:
                        continue
                    rd = {headers[i]: str(v) for i, v in enumerate(row) if i < len(headers) and headers[i] and v}
                    if rd:
                        rows.append(rd)
            wb.close()
        except Exception as exc:
            logger.warning("_collect_example_rows falhou: %s", exc)
        return rows

    # ── Privadas ──────────────────────────────────────────────────────────────

    def _read_template_headers(
        self, template_bytes: bytes, marketplace: str
    ) -> Optional[dict[int, str]]:
        import os
        import tempfile
        from openpyxl import load_workbook

        _INLINE_CONFIG: dict[str, dict] = {
            "Mercado Livre": {"sheet_index_after_ajuda": True, "header_row": 3},
        }

        try:
            from core.filler import MARKETPLACE_CONFIG
            config = MARKETPLACE_CONFIG.get(marketplace) or _INLINE_CONFIG.get(marketplace)
        except Exception:
            config = _INLINE_CONFIG.get(marketplace)

        if not config:
            logger.error(
                "Marketplace '%s' não encontrado em MARKETPLACE_CONFIG.",
                marketplace,
            )
            return None

        tmp_path = None
        sanitized_path = None
        try:
            from core.xlsx_openpyxl_compat import sanitize_xlsx_for_openpyxl

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(template_bytes)
                tmp_path = tmp.name

            sanitized_path = sanitize_xlsx_for_openpyxl(tmp_path)
            load_path = sanitized_path or tmp_path

            # read_only=True breaks multi-column iteration on some xlsx files (e.g. Shopee)
            wb = load_workbook(load_path)
            logger.info("Template '%s' — abas: %s", marketplace, wb.sheetnames)

            if marketplace == "Vendor":
                prefix = config.get("sheet_prefix", "Modelo-")
                ws = next(
                    (wb[s] for s in wb.sheetnames if s.startswith(prefix)),
                    wb.active,
                )
            elif marketplace == "Amazon":
                # Amazon como destino: tenta candidatos em ordem
                candidates = config.get("sheet_candidates", ["Template", "Modelo"])
                ws = None
                for cand in candidates:
                    if cand in wb.sheetnames:
                        ws = wb[cand]
                        break
                if ws is None:
                    for cand in candidates:
                        for sn in wb.sheetnames:
                            if cand.lower() in sn.lower():
                                ws = wb[sn]
                                break
                        if ws:
                            break
                if ws is None:
                    ws = wb.active
            elif config.get("sheet_index_after_ajuda"):
                if len(wb.sheetnames) >= 3:
                    ws = wb[wb.sheetnames[2]]
                else:
                    ws = wb[wb.sheetnames[-1]]
            else:
                sheet = config.get("sheet", "")
                if not sheet:
                    candidates = config.get("sheet_candidates", [])
                    sheet = next((s for s in candidates if s in wb.sheetnames), "")
                    if not sheet:
                        for cand in candidates:
                            for sn in wb.sheetnames:
                                if cand.lower() in sn.lower():
                                    sheet = sn
                                    break
                            if sheet:
                                break
                    if not sheet:
                        _skip = {"instruções","instrucoes","instructions","ajuda","help","dropdown","conditions","cover"}
                        sheet = next(
                            (sn for sn in wb.sheetnames if not any(sk in sn.lower() for sk in _skip)),
                            wb.sheetnames[0] if wb.sheetnames else ""
                        )
                ws = wb[sheet] if sheet in wb.sheetnames else wb.active

            zp = _find_sheet_zip_path(template_bytes, ws.title)
            header_row = config["header_row"]
            if zp:
                try:
                    with zipfile.ZipFile(io.BytesIO(template_bytes), "r") as zf:
                        header_row = guess_best_header_row_from_sheet_xml(
                            zf.read(zp), config["header_row"]
                        )
                except Exception as exc:
                    logger.warning(
                        "Cabeçalho via XML falhou (%s); tentando openpyxl.", exc,
                    )
                    header_row = guess_best_header_row(ws, config["header_row"])
            else:
                logger.warning(
                    "Não foi possível localizar a aba '%s' no ZIP (workbook); "
                    "cabeçalho só via openpyxl.",
                    ws.title,
                )
                header_row = guess_best_header_row(ws, config["header_row"])

            headers = {}
            for cell in ws[header_row]:
                if cell.value:
                    headers[cell.column] = cell.value

            logger.info("Cabeçalhos lidos: %d colunas na aba '%s'", len(headers), ws.title)
            wb.close()
            return headers

        except Exception as exc:
            logger.error(
                "Erro ao ler cabeçalhos do template '%s': %s", marketplace, exc,
                exc_info=True,
            )
            return None
        finally:
            for _p in (sanitized_path, tmp_path):
                if _p:
                    try:
                        os.unlink(_p)
                    except OSError:
                        pass

    def _apply_enrichment(
        self, df: pd.DataFrame, marketplace: str
    ) -> pd.DataFrame:
        """Aplica enriquecimento de IA linha por linha (com cache automático)."""
        enriched_rows = []
        for _, row in df.iterrows():
            row_dict = row.dropna().to_dict()
            enrich = self._ai.enrich_row(row_dict, marketplace)
            if enrich:
                for col_alias, field_key in [
                    ("item name", "title"),
                    ("nome do produto", "title"),
                    ("nome_produto", "title"),
                    ("product description", "description"),
                    ("descrição do produto", "description"),
                    ("descricao", "description"),
                ]:
                    if col_alias in row.index and field_key in enrich:
                        row[col_alias] = enrich[field_key]
                bullets = enrich.get("bullets", [])
                for i, bullet in enumerate(bullets[:5], start=1):
                    col_bp = f"bullet point{i}" if i > 1 else "bullet point"
                    if col_bp in row.index:
                        row[col_bp] = bullet
            enriched_rows.append(row)

        return pd.DataFrame(enriched_rows)