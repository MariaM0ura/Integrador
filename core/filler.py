"""
filler.py
=========
Preenchimento de templates de marketplaces.

ARQUITETURA:
  Em vez de usar openpyxl para abrir+salvar o arquivo (o que destrói o
  Content-Type, as x14:dataValidations e outros metadados), este módulo
  usa uma abordagem cirúrgica via zipfile:

  1. Copia TODOS os arquivos do ZIP original sem modificação.
  2. Apenas o XML da worksheet de dados é modificado — os valores das
     células são inseridos/sobrescritos diretamente no XML.
  3. Todo o resto (styles.xml, [Content_Types].xml, x14:dataValidations,
     Dropdown Lists, fórmulas INDIRECT, etc.) é preservado bit-a-bit.

  Isso resolve:
  - Arquivo corrompido: o Content-Type macroEnabled é preservado.
  - Dropdowns perdidos: as x14:dataValidations (INDIRECT) são preservadas.
  - Formatação: estilos, bordas e cores não são tocados.
"""

from __future__ import annotations

import copy
import io
import logging
import re
import tempfile
import unicodedata
import uuid
from datetime import datetime
import warnings
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd

from core.mapper import MappingResult, REQUIRED_FIELDS
from core.normalizer import FieldNormalizer

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
logger = logging.getLogger(__name__)


# ─── Configuração por marketplace ─────────────────────────────────────────────

MARKETPLACE_CONFIG: dict[str, dict] = {
    "Temu": {
        "sheet": "Template",
        "header_row": 2,
        "data_start": 5,
    },
    "Shopee": {
        "sheet": "Modelo",
        "sheet_candidates": ["Modelo", "MODELO", "Modelo de produto", "Product template"],
        # Muitos templates PT têm metadados nas linhas 1–2; cabeçalhos reais podem ser 2–5.
        "header_row": 3,
        "data_start": 7,
    },
    "Vendor": {
        "sheet_prefix": "Modelo-",
        "header_row": 3,
        "data_start": 9,
    },
    "Mercado Livre": {
        "sheet_index_after_ajuda": True,
        "header_row": 3,
        "data_start": 9,
    },
    "Magalu": {
        "sheet": "PRODUTO",
        "header_row": 3,
        "data_start": 5,
        "unit_conversions": {
            # Somente Magalu usa metros — converter dimensões cm→m
            "altura_pacote":      ("cm", "m"),
            "largura_pacote":     ("cm", "m"),
            "comprimento_pacote": ("cm", "m"),
        },
    },
    "Walmart": {
        "sheet": "Product Content And Site Exp",
        "header_row": 4,
        "data_start": 7,
    },
    "Amazon": {
        # Amazon como DESTINO: recebe dados de outros marketplaces.
        # Origens em kg/cm → Amazon espera lb/in.
        "sheet": "Template",
        "sheet_candidates": ["Template", "Modelo"],
        "header_row": 4,
        "data_start": 7,
    },
}

FIELD_TYPE_HINTS: dict[str, list[str]] = {
    "cor": ["cor", "color", "colour"],
    "tamanho": ["tamanho", "size"],
    "preco": ["preço", "preco", "price", "base price"],
    "peso_pacote": ["peso", "weight"],
    "comprimento_pacote": ["comprimento", "length", "profundidade"],
    "largura_pacote": ["largura", "width"],
    "altura_pacote": ["altura", "height"],
}


# ─── Dataclasses ──────────────────────────────────────────────────────────────

@dataclass
class ValidationIssue:
    column: str
    issue_type: str
    severity: str
    message: str


@dataclass
class FillResult:
    output_path: Optional[str]
    marketplace: str
    rows_written: int
    validation_issues: list[ValidationIssue] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return not any(i.severity == "error" for i in self.validation_issues)

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)


# ─── Helpers de coluna ────────────────────────────────────────────────────────

def _col_letter(n: int) -> str:
    """Converte índice 1-based para letra(s) Excel (1→A, 26→Z, 27→AA)."""
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def _strip_accents(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )


def _normalize_col(text) -> str:
    return _strip_accents(str(text).strip().lower())


def guess_best_header_row(ws, default_row: int, *, min_cols: int = 5, scan_to: int = 15) -> int:
    """
    Alguns templates (ex.: Shopee) mudam a linha dos títulos entre versões/idiomas.
    Se a linha configurada tiver poucas células preenchidas, escolhe a linha em
    1..scan_to com o maior número de células não vazias.
    """
    def _count_nonempty(row_idx: int) -> int:
        n = 0
        for cell in ws[row_idx]:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str) and not str(v).strip():
                continue
            n += 1
        return n

    try:
        base = _count_nonempty(default_row)
    except Exception:
        return default_row

    if base >= min_cols:
        return default_row

    best_row, best_cnt = default_row, base
    for r in range(1, scan_to + 1):
        try:
            c = _count_nonempty(r)
        except Exception:
            continue
        if c > best_cnt:
            best_cnt, best_row = c, r

    if best_row != default_row:
        logger.info(
            "guess_best_header_row: aba '%s' — linha %s tinha %s col.; usando linha %s (%s col.)",
            getattr(ws, "title", "?"),
            default_row,
            base,
            best_row,
            best_cnt,
        )
    return best_row


def compute_data_start_row(config: dict, resolved_header_row: int) -> int:
    """Mantém o deslocamento header→dados do config, sem subir acima da linha antiga de dados."""
    orig_h = int(config["header_row"])
    orig_d = int(config["data_start"])
    delta = max(1, orig_d - orig_h)
    return max(orig_d, resolved_header_row + delta)


def _norm_sheet_title(s: str) -> str:
    return " ".join(s.replace("\xa0", " ").strip().split()).casefold()


def guess_best_header_row_from_sheet_xml(
    sheet_xml: bytes,
    default_row: int,
    *,
    min_cols: int = 5,
    scan_to: int = 40,
) -> int:
    """
    Infere a linha de cabeçalho contando células (refs A1, B2, …) no XML da aba.
    Mais fiel que openpyxl read-only para templates com muitas células vazias na
    linha configurada (ex.: Shopee).
    """
    text = sheet_xml.decode("utf-8", errors="replace")
    counts: dict[int, int] = {}
    for _col, rn_s in re.findall(
        r'<c\b[^>]*\br="([A-Z]{1,3})(\d+)"', text, flags=re.IGNORECASE
    ):
        try:
            rn = int(rn_s)
        except ValueError:
            continue
        if rn < 1 or rn > scan_to:
            continue
        counts[rn] = counts.get(rn, 0) + 1

    if not counts:
        return default_row

    default_cnt = counts.get(default_row, 0)
    if default_cnt >= min_cols:
        return default_row

    best_r, best_cnt = default_row, default_cnt
    for r, c in counts.items():
        if r > scan_to:
            continue
        if c > best_cnt or (c == best_cnt and r < best_r):
            best_r, best_cnt = r, c

    if best_cnt > default_cnt:
        logger.info(
            "guess_best_header_row_from_sheet_xml: linha padrão %s (%s <c>) → linha %s (%s <c>)",
            default_row,
            default_cnt,
            best_r,
            best_cnt,
        )
        return best_r
    return default_row


# ─── Escrita cirúrgica no XML ─────────────────────────────────────────────────

def _inject_values_into_sheet_xml(
    sheet_xml_bytes: bytes,
    data_start_row: int,
    row_col_values: dict[int, dict[int, object]],
) -> bytes:
    """
    Insere valores nas células de uma worksheet XML sem tocar em nada mais.

    Usa manipulação de string pura — NÃO faz parse/serialize do XML completo,
    o que destruiria namespaces customizados (x14, x14ac, etc.) e as
    x14:dataValidations (dropdowns condicionais).

    Estratégia:
      - Para cada linha de dados, gera um fragmento <row>...</row> como string.
      - Se a row já existe no XML: substitui via regex.
      - Se não existe: insere antes de </sheetData>.
      - Tudo fora do sheetData é preservado byte-a-bit.
    """
    sheet_xml = sheet_xml_bytes.decode("utf-8")

    for row_offset, col_vals in row_col_values.items():
        row_num = data_start_row + row_offset
        new_row = _build_row_xml(row_num, col_vals)

        existing = re.search(rf'<row r="{row_num}"[^>]*>.*?</row>', sheet_xml, re.DOTALL)
        if existing:
            et = existing.group(0)
            rm = re.match(r'<row(\b[^>]*)>', et)
            orig_row_attrs = rm.group(1) if rm else ""
            orig_cells = {}
            for cm in re.finditer(r'(<c\b[^>]*?/>|<c\b[^>]*>.*?</c>)', et, re.DOTALL):
                ref_m = re.search(r'\br="([^"]+)"', cm.group(0))
                if ref_m:
                    orig_cells[ref_m.group(1)] = cm.group(0)
            new_row = _build_row_xml(row_num, col_vals, orig_row_attrs, orig_cells)
            sheet_xml = sheet_xml[:existing.start()] + new_row + sheet_xml[existing.end():]
        else:
            insert_before = None
            for m in re.finditer(r'<row r="(\d+)"[^>]*>', sheet_xml):
                if int(m.group(1)) > row_num:
                    insert_before = m.start()
                    break

            if insert_before is not None:
                sheet_xml = sheet_xml[:insert_before] + new_row + sheet_xml[insert_before:]
            else:
                sheet_xml = sheet_xml.replace("</sheetData>", new_row + "</sheetData>", 1)

    return sheet_xml.encode("utf-8")


def _build_row_xml(
    row_num: int, col_vals: dict[int, object],
    orig_row_attrs: str = "", orig_cells: "dict[str, str] | None" = None,
) -> str:
    """Gera XML de uma <row> preservando atributos e estilos originais."""
    if orig_row_attrs.strip():
        attrs = re.sub(r'\br="[^"]*"\s*', '', orig_row_attrs).strip()
        row_open = f'<row r="{row_num}" {attrs}>' if attrs else f'<row r="{row_num}">'
    else:
        row_open = f'<row r="{row_num}">'
    new_refs: dict[str, str] = {}
    for col_idx, value in sorted(col_vals.items()):
        if value is None or str(value).strip() in ("", "nan", "None"):
            continue
        ref = f"{_col_letter(col_idx)}{row_num}"
        style = ""
        if orig_cells and ref in orig_cells:
            s_m = re.search(r'\bs="([^"]+)"', orig_cells[ref])
            if s_m:
                style = f' s="{s_m.group(1)}"'
        if isinstance(value, str):
            v = value.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
            # Usar t="str" em vez de t="inlineStr": compatível com data validation
            # lists (dropdowns). O inlineStr faz o Excel ignorar o dropdown da célula.
            new_refs[ref] = f'<c r="{ref}"{style} t="str"><v>{v}</v></c>'
        else:
            new_refs[ref] = f'<c r="{ref}"{style}><v>{value}</v></c>'
    all_cells: dict[str, str] = dict(orig_cells) if orig_cells else {}
    all_cells.update(new_refs)

    def _col_key(ref):
        m = re.match(r'([A-Z]+)(\d+)', ref)
        if not m:
            return (0, 0)
        n = sum((ord(ch) - 64) * (26 ** i) for i, ch in enumerate(reversed(m.group(1))))
        return (int(m.group(2)), n)

    return f'{row_open}{"".join(all_cells[r] for r in sorted(all_cells, key=_col_key))}</row>'


# ─── Mapeamento de sheet name → arquivo no ZIP ───────────────────────────────

def _find_sheet_zip_path(
    template_bytes: bytes, sheet_name: str
) -> Optional[str]:
    """
    Retorna o caminho ZIP da worksheet (ex.: xl/worksheets/sheet3.xml).

    Aceita ``<sheet>`` com ou sem auto-fechamento, prefixo de namespace e
    ``Target`` com ou sem barra inicial; compara nomes de aba sem distinção
    de maiúsculas e com espaços normalizados.
    """
    target_cf = _norm_sheet_title(sheet_name)

    def _resolve_target_in_zip(z: zipfile.ZipFile, target_raw: str) -> str:
        tgt = target_raw.replace("\\", "/").strip()
        if tgt.startswith("/"):
            tgt = tgt.lstrip("/")
        if not tgt.startswith("xl/"):
            tgt = "xl/" + tgt.lstrip("/")
        norm = {n.replace("\\", "/").lower(): n for n in z.namelist()}
        key = tgt.lower()
        return norm.get(key, tgt)

    buf = io.BytesIO(template_bytes)
    with zipfile.ZipFile(buf, "r") as z:
        try:
            wb_xml = z.read("xl/workbook.xml").decode("utf-8", errors="replace")
        except KeyError:
            logger.warning("_find_sheet_zip_path: xl/workbook.xml ausente")
            return None

        sheets: list[tuple[str, str]] = []
        for m in re.finditer(
            r"<(?:[\w.\-]+:)?sheet\b([^>]+)>",
            wb_xml,
            flags=re.IGNORECASE | re.DOTALL,
        ):
            chunk = m.group(1).strip()
            if chunk.endswith("/"):
                chunk = chunk[:-1].strip()
            name_m = re.search(r'name="([^"]*)"', chunk, flags=re.IGNORECASE)
            rid_m = re.search(r'r:id="([^"]*)"', chunk, flags=re.IGNORECASE)
            if not name_m or not rid_m:
                continue
            sheets.append((name_m.group(1), rid_m.group(1)))

        if not sheets:
            return None

        rid: Optional[str] = None
        for raw, r in sheets:
            if raw == sheet_name or _norm_sheet_title(raw) == target_cf:
                rid = r
                break
        if rid is None:
            for raw, r in sheets:
                r_cf = _norm_sheet_title(raw)
                if target_cf in r_cf or r_cf in target_cf:
                    rid = r
                    break

        if rid is None:
            logger.warning(
                "_find_sheet_zip_path: aba '%s' não encontrada em workbook.xml (ex.: %s)",
                sheet_name,
                [s[0] for s in sheets[:10]],
            )
            return None

        try:
            rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="replace")
        except KeyError:
            logger.warning("_find_sheet_zip_path: xl/_rels/workbook.xml.rels ausente")
            return None

        t_m = re.search(
            rf'Id="{re.escape(rid)}"[^>]*Target="([^"]+)"',
            rels_xml,
            flags=re.IGNORECASE | re.DOTALL,
        )
        if not t_m:
            t_m = re.search(
                rf'Target="([^"]+)"[^>]*Id="{re.escape(rid)}"',
                rels_xml,
                flags=re.IGNORECASE | re.DOTALL,
            )
        if not t_m:
            logger.warning(
                "_find_sheet_zip_path: relação Id=%s não encontrada em workbook.xml.rels",
                rid,
            )
            return None

        return _resolve_target_in_zip(z, t_m.group(1))


# ─── Filler ───────────────────────────────────────────────────────────────────

class MarketplaceFiller:
    """
    Preenche templates Excel de marketplaces com dados de origem.

    Usa abordagem cirúrgica via zipfile: preserva TUDO do arquivo original
    (Content-Type, x14:dataValidations, estilos, fórmulas INDIRECT, etc.)
    e apenas injeta os valores das células no XML da worksheet de dados.

    Suporta qualquer marketplace como destino, inclusive Amazon.
    """

    def __init__(self):
        self._normalizer = FieldNormalizer()

    def _resolve_layout_rows(
        self, template_bytes: bytes, sheet_name: str, config: dict
    ) -> tuple[int, int]:
        """Infere linha de cabeçalho + início dos dados (XML da aba; fallback openpyxl)."""
        import os
        import tempfile

        from openpyxl import load_workbook

        from core.xlsx_openpyxl_compat import sanitize_xlsx_for_openpyxl

        zp = _find_sheet_zip_path(template_bytes, sheet_name)
        if zp:
            try:
                with zipfile.ZipFile(io.BytesIO(template_bytes), "r") as zf:
                    sxml = zf.read(zp)
                hr = guess_best_header_row_from_sheet_xml(
                    sxml, config["header_row"]
                )
                return hr, compute_data_start_row(config, hr)
            except Exception as exc:
                logger.warning("layout via XML (%s) falhou: %s", zp, exc)

        tmp_path = None
        sanitized_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(template_bytes)
                tmp_path = tmp.name
            sanitized_path = sanitize_xlsx_for_openpyxl(tmp_path)
            load_path = sanitized_path or tmp_path
            wb = load_workbook(load_path, read_only=True)
            try:
                ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
                hr = guess_best_header_row(ws, config["header_row"])
                dr = compute_data_start_row(config, hr)
                return hr, dr
            finally:
                wb.close()
        finally:
            for _p in (sanitized_path, tmp_path):
                if _p:
                    try:
                        os.unlink(_p)
                    except OSError:
                        pass

    # ── Pública ───────────────────────────────────────────────────────────────

    def fill(
        self,
        amazon_df: pd.DataFrame,
        mapping: MappingResult,
        template_file,
        output_dir: Optional[str] = None,
        template_ext: Optional[str] = None,
    ) -> FillResult:
        marketplace = mapping.marketplace
        config = MARKETPLACE_CONFIG.get(marketplace)
        if not config:
            return FillResult(
                output_path=None, marketplace=marketplace, rows_written=0,
                errors=[f"Marketplace '{marketplace}' não configurado em MARKETPLACE_CONFIG."],
            )

        # ── Ler template como bytes ───────────────────────────────────────
        try:
            if hasattr(template_file, "seek"):
                template_file.seek(0)
                template_bytes = template_file.read()
            else:
                template_bytes = Path(template_file).read_bytes()

            if template_ext and template_ext in (".xlsx", ".xlsm"):
                original_ext = template_ext
            elif hasattr(template_file, "name") and template_file.name:
                original_ext = Path(template_file.name).suffix.lower() or ".xlsx"
            else:
                original_ext = ".xlsx"
                try:
                    import io as _io2
                    with zipfile.ZipFile(_io2.BytesIO(template_bytes)) as _zchk:
                        if "xl/vbaProject.bin" in _zchk.namelist():
                            original_ext = ".xlsm"
                except Exception:
                    pass
            if original_ext not in (".xlsx", ".xlsm"):
                original_ext = ".xlsx"
        except Exception as exc:
            return FillResult(
                output_path=None, marketplace=marketplace, rows_written=0,
                errors=[f"Erro ao ler template: {exc}"],
            )

        # ── Resolver nome da aba e caminho no ZIP ─────────────────────────
        sheet_name = self._resolve_sheet_name(template_bytes, config, marketplace)
        if sheet_name is None:
            return FillResult(
                output_path=None, marketplace=marketplace, rows_written=0,
                errors=["Aba do template não encontrada."],
            )

        sheet_zip_path = _find_sheet_zip_path(template_bytes, sheet_name)
        if sheet_zip_path is None:
            return FillResult(
                output_path=None, marketplace=marketplace, rows_written=0,
                errors=[f"Não foi possível localizar '{sheet_name}' no ZIP."],
            )

        try:
            header_row, data_start = self._resolve_layout_rows(
                template_bytes, sheet_name, config
            )
        except Exception as exc:
            logger.warning(
                "Não foi possível inferir layout do template (%s); usando MARKETPLACE_CONFIG.",
                exc,
            )
            header_row = int(config["header_row"])
            data_start = int(config["data_start"])

        # ── Detectar tipos de campo pelos headers ─────────────────────────
        field_types = self._detect_field_types_from_zip(
            template_bytes, sheet_zip_path, header_row
        )

        # ── Construir dicionário de valores por linha/coluna ──────────────
        unit_conversions = config.get("unit_conversions", {})
        row_col_values: dict[int, dict[int, object]] = {}

        for row_offset, (_, amazon_row) in enumerate(amazon_df.iterrows()):
            col_vals: dict[int, object] = {}
            for col_idx, source_idx in mapping.index_map.items():
                raw_value = (
                    amazon_row.iloc[source_idx]
                    if source_idx < len(amazon_row)
                    else None
                )
                field_type = field_types.get(col_idx, "")
                value = self._normalize_value(raw_value, field_type, unit_conversions)
                col_vals[col_idx] = value
            row_col_values[row_offset] = col_vals

        # ── Escrever via zipfile cirúrgico ────────────────────────────────
        rows_written = len(row_col_values)
        out_dir = Path(output_dir) if output_dir else Path(tempfile.gettempdir())
        out_dir.mkdir(parents=True, exist_ok=True)
        _ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
        _uid = uuid.uuid4().hex[:6]
        output_path = str(out_dir / f"{marketplace}_preenchido_{_ts}_{_uid}{original_ext}")

        try:
            import io
            in_buf = io.BytesIO(template_bytes)
            out_buf = io.BytesIO()

            with zipfile.ZipFile(in_buf, "r") as zin, \
                 zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename == sheet_zip_path:
                        data = _inject_values_into_sheet_xml(
                            data, data_start, row_col_values
                        )
                    zout.writestr(item, data)

            with open(output_path, "wb") as f:
                f.write(out_buf.getvalue())

        except Exception as exc:
            return FillResult(
                output_path=None, marketplace=marketplace,
                rows_written=rows_written,
                errors=[f"Erro ao salvar arquivo de saída: {exc}"],
            )

        logger.info(
            "Template %s preenchido: %d linhas → %s",
            marketplace, rows_written, output_path,
        )
        return FillResult(
            output_path=output_path,
            marketplace=marketplace,
            rows_written=rows_written,
            validation_issues=[],
        )

    # ── Privadas ──────────────────────────────────────────────────────────────

    @staticmethod
    def _resolve_sheet_name(
        template_bytes: bytes, config: dict, marketplace: str
    ) -> Optional[str]:
        """Encontra o nome da aba correta dentro do ZIP."""
        with zipfile.ZipFile(io.BytesIO(template_bytes)) as z:
            wb_xml = z.read("xl/workbook.xml").decode("utf-8", errors="replace")
        sheet_names = re.findall(
            r'<(?:[\w.\-]+:)?sheet[^>]*name="([^"]+)"', wb_xml, flags=re.IGNORECASE
        )

        def _candidates() -> list[str]:
            out: list[str] = []
            seen: set[str] = set()
            for key in ("sheet",):
                v = config.get(key)
                if isinstance(v, str) and v.strip() and v not in seen:
                    seen.add(v)
                    out.append(v)
            for v in config.get("sheet_candidates") or []:
                if isinstance(v, str) and v.strip() and v not in seen:
                    seen.add(v)
                    out.append(v)
            return out

        # Amazon como destino: tenta candidatos em ordem
        if marketplace == "Amazon":
            candidates = config.get("sheet_candidates", ["Template", "Modelo"])
            for cand in candidates:
                if cand in sheet_names:
                    return cand
            for cand in candidates:
                for s in sheet_names:
                    if cand.lower() in s.lower():
                        return s
            return sheet_names[0] if sheet_names else None

        if marketplace == "Vendor":
            prefix = config.get("sheet_prefix", "Modelo-")
            for name in sheet_names:
                if name.startswith(prefix):
                    return name
            return sheet_names[0] if sheet_names else None

        # Mercado Livre: aba de dados é sempre a terceira aba (índice 2).
        if config.get("sheet_index_after_ajuda"):
            if len(sheet_names) >= 3:
                target_name = sheet_names[2]
                logger.info(
                    "Mercado Livre: aba resolvida como '%s' (terceira aba, índice 2).",
                    target_name,
                )
                return target_name
            target_name = sheet_names[-1] if sheet_names else None
            logger.warning("Mercado Livre: estrutura inesperada; usando '%s'.", target_name)
            return target_name

        cand_list = _candidates()
        for cand in cand_list:
            if cand in sheet_names:
                return cand
        for cand in cand_list:
            c_low = cand.lower()
            for s in sheet_names:
                if c_low in s.lower() or s.lower() in c_low:
                    logger.warning("Aba candidata '%s' → '%s'.", cand, s)
                    return s

        target = config.get("sheet", "")
        if target in sheet_names:
            return target
        for name in sheet_names:
            if target and target.lower() in name.lower():
                logger.warning("Aba '%s' não encontrada; usando '%s'.", target, name)
                return name
        return sheet_names[0] if sheet_names else None

    def _detect_field_types_from_zip(
        self, template_bytes: bytes, sheet_zip_path: str, header_row: int
    ) -> dict[int, str]:
        """Detecta tipos de campo lendo o XML do sheet diretamente."""
        import io
        try:
            with zipfile.ZipFile(io.BytesIO(template_bytes)) as z:
                sheet_xml = z.read(sheet_zip_path)  # noqa: F841 — reservado para uso futuro

            from openpyxl import load_workbook

            from core.xlsx_openpyxl_compat import sanitize_xlsx_for_openpyxl

            tmp_path = None
            sanitized_path = None
            try:
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    tmp.write(template_bytes)
                    tmp_path = tmp.name
                sanitized_path = sanitize_xlsx_for_openpyxl(tmp_path)
                load_path = sanitized_path or tmp_path
                wb = load_workbook(load_path, read_only=True)
                types: dict[int, str] = {}
                for ws_name in wb.sheetnames:
                    ws = wb[ws_name]
                    hr = guess_best_header_row(ws, header_row)
                    for cell in ws[hr]:
                        if not cell.value:
                            continue
                        col_name = _normalize_col(cell.value)
                        for field_type, hints in FIELD_TYPE_HINTS.items():
                            if any(h in col_name for h in hints):
                                types[cell.column] = field_type
                                break
                    if types:
                        break
                wb.close()
                return types
            finally:
                import os as _os

                for _p in (sanitized_path, tmp_path):
                    if _p:
                        try:
                            _os.unlink(_p)
                        except OSError:
                            pass
        except Exception as exc:
            logger.warning("Não foi possível detectar tipos de campo: %s", exc)
            return {}

    # Campos que devem ser arredondados para 2 casas decimais na saída
    _TWO_DECIMAL_FIELDS = {
        "peso_pacote", "comprimento_pacote", "largura_pacote", "altura_pacote",
    }

    def _normalize_value(self, value, field_type: str, unit_conversions: dict):
        if pd.isna(value) or str(value).strip() in ("", "nan", "None"):
            return None
        if not field_type:
            return value
        conv = unit_conversions.get(field_type)
        kwargs = {}
        if conv:
            kwargs = {"from_unit": conv[0], "to_unit": conv[1]}
        result = self._normalizer.normalize_field(field_type, value, **kwargs)
        normalized = result.normalized if result.normalized != "" else None
        if normalized is not None and field_type in self._TWO_DECIMAL_FIELDS:
            try:
                normalized = round(float(normalized), 2)
            except (ValueError, TypeError):
                pass
        return normalized

    def _validate_output(
        self, output_path: str, sheet_name: str, config: dict, marketplace: str
    ) -> list[ValidationIssue]:
        """Valida campos obrigatórios lendo o arquivo gerado."""
        issues: list[ValidationIssue] = []
        required = REQUIRED_FIELDS.get(marketplace, [])
        if not required:
            return issues

        try:
            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

            hr = guess_best_header_row(ws, config["header_row"])
            data_start = compute_data_start_row(config, hr)

            col_map: dict[str, int] = {}
            for cell in ws[hr]:
                if cell.value:
                    key = str(cell.value).strip().lower()
                    if key not in col_map:
                        col_map[key] = cell.column

            last_data_row = data_start - 1
            for row in ws.iter_rows(min_row=data_start):
                if any(c.value is not None and str(c.value).strip() != "" for c in row):
                    last_data_row = row[0].row
            if last_data_row < data_start:
                return issues

            for req in required:
                req_norm = str(req).strip().lower()
                if req_norm not in col_map:
                    issues.append(ValidationIssue(
                        column=req, issue_type="missing_required", severity="warning",
                        message=f"Campo obrigatório '{req}' não encontrado no template.",
                    ))
                    continue

                col_idx = col_map[req_norm]
                empty_rows = []
                for row in ws.iter_rows(
                    min_row=data_start, max_row=last_data_row,
                    min_col=col_idx, max_col=col_idx
                ):
                    cell = row[0]
                    if cell.value is None or str(cell.value).strip() == "":
                        empty_rows.append(cell.row)

                if empty_rows:
                    issues.append(ValidationIssue(
                        column=req, issue_type="empty_value", severity="error",
                        message=(
                            f"Campo obrigatório '{req}' vazio nas linhas: "
                            f"{', '.join(str(r) for r in empty_rows[:10])}"
                            f"{'...' if len(empty_rows) > 10 else ''}."
                        ),
                    ))
        except Exception as exc:
            logger.warning("Erro na validação do output: %s", exc)

        return issues
