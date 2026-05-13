"""
instruction_parser.py
=====================
Extrai regras de preenchimento das abas de instrução dos templates de marketplace.

Suporta dois formatos de instrução:
  - Inline: linhas de metadados dentro da aba de dados (TikTok, Shopee)
    → detecta linhas "Obrigatório/Opcional" e linhas de descrição por coluna
  - Tabela separada: aba de definições com mapeamento col→regra (Amazon, Walmart)
    → parseia sheets com nomes como "Definições de dados", "Data Definitions"
  - Exemplo: extrai valores de abas de exemplo para preenchimento por herança
"""

from __future__ import annotations

import logging
import re
import unicodedata
from dataclasses import dataclass, field
from typing import Optional

logger = logging.getLogger(__name__)

# ─── Padrões de nome de aba ───────────────────────────────────────────────────

_INSTRUCTION_PATTERNS = [
    "instrucao", "instrucoes", "instruction", "instructions",
    "definicoes de dados", "data definitions", "data definition",
    "orientacao", "ajuda", "help", "como preencher",
    "definicao", "definition",
]

_EXAMPLE_PATTERNS = [
    "exemplo", "exemplos", "example", "examples",
    "fazer upload do exemplo", "sample", "amostras",
]

_REQUIRED_KW = {"obrigatório", "obrigatorio", "required", "requerido", "sim"}
_OPTIONAL_KW = {"opcional", "optional", "nao obrigatorio", "não obrigatório", "não"}
# Regex for Shopee/TikTok internal column codes: "ps_product_name|1|0", "et_title_variation_1|0|0"
_INTERNAL_CODE_RE = re.compile(r"^[a-z0-9_]+(\|[^|]+){1,4}$")


# ─── Dataclass de regra por coluna ────────────────────────────────────────────

@dataclass
class ColumnRule:
    nome_coluna: str
    regra: str = ""
    valores_aceitos: list[str] = field(default_factory=list)
    obrigatorio: bool = False
    exemplo: str = ""
    tipo: str = "text"          # text | number | date | list


# ─── Parser ───────────────────────────────────────────────────────────────────

class InstructionParser:
    """
    Parseia abas de instrução de templates xlsx para extrair regras por coluna.

    Uso:
        parser = InstructionParser()
        rules = parser.parse(template_bytes, marketplace, sheet_name, header_row)
        # → {col_name: ColumnRule}
    """

    def parse(
        self,
        template_bytes: bytes,
        marketplace: str,
        data_sheet_name: str,
        header_row: int = 3,
    ) -> dict[str, ColumnRule]:
        """
        Retorna {nome_coluna: ColumnRule} para todas as colunas do template.
        Combina múltiplas fontes de instrução sem duplicar entradas.
        """
        rules: dict[str, ColumnRule] = {}
        try:
            wb = _load_wb(template_bytes)

            # 1. Extrai metadados inline da aba de dados
            if data_sheet_name in wb.sheetnames:
                ws = wb[data_sheet_name]
                inline = self._parse_inline_metadata(ws, header_row)
                rules.update(inline)

            # 2. Enriquece com abas de definição separadas
            for sheet_name in wb.sheetnames:
                sn = _norm(sheet_name)
                if any(p in sn for p in _INSTRUCTION_PATTERNS):
                    ws = wb[sheet_name]
                    extra = self._parse_definition_sheet(ws, rules)
                    for col, rule in extra.items():
                        if col in rules:
                            # Merge: atualiza campos vazios
                            _merge_rule(rules[col], rule)
                        else:
                            rules[col] = rule

            # 3. Extrai exemplos de abas de exemplo
            for sheet_name in wb.sheetnames:
                sn = _norm(sheet_name)
                if any(p in sn for p in _EXAMPLE_PATTERNS):
                    ws = wb[sheet_name]
                    self._extract_examples(ws, rules)

            wb.close()
        except Exception as exc:
            logger.warning("InstructionParser falhou para '%s': %s", marketplace, exc)

        logger.info(
            "InstructionParser '%s': %d regras extraídas (%d obrigatórias)",
            marketplace,
            len(rules),
            sum(1 for r in rules.values() if r.obrigatorio),
        )
        return rules

    # ── Inline metadata ───────────────────────────────────────────────────────

    def _parse_inline_metadata(
        self, ws, header_row: int
    ) -> dict[str, ColumnRule]:
        """
        Detecta linhas de metadados inline na aba de dados.
        Padrão Shopee/TikTok: rows antes/em torno do header_row contém
        required/optional flags e descrições por coluna.
        """
        rules: dict[str, ColumnRule] = {}
        scan_limit = min(header_row + 4, 15)

        # Materializa as primeiras linhas
        rows_data: dict[int, list] = {}
        for r_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=scan_limit, values_only=True),
            start=1,
        ):
            rows_data[r_idx] = list(row)

        # Trust header_row param if it has content; only auto-detect as fallback
        best_row = header_row
        header_data = rows_data.get(header_row, [])
        if sum(1 for v in header_data if v and str(v).strip()) < 3:
            best_cnt = 0
            for r_idx, row in rows_data.items():
                cnt = sum(1 for v in row if v and str(v).strip())
                if cnt > best_cnt:
                    best_cnt, best_row = cnt, r_idx
            header_data = rows_data.get(best_row, [])

        # Mapeia col_idx → nome_coluna
        col_names: dict[int, str] = {}
        for col_idx, val in enumerate(header_data, start=1):
            if val and str(val).strip():
                col_names[col_idx] = str(val).strip()

        # Para cada coluna, varre as outras linhas buscando required/desc
        for col_idx, col_name in col_names.items():
            rule = ColumnRule(nome_coluna=col_name)
            for r_idx, row in rows_data.items():
                if r_idx == best_row:
                    continue
                if col_idx > len(row):
                    continue
                cell_val = row[col_idx - 1]
                if cell_val is None:
                    continue
                cell_str = str(cell_val).strip()
                cell_norm = _norm(cell_str)

                # "Condicional obrigatório" ≠ truly mandatory — skip it
                is_conditional = "condicional" in cell_norm
                if not is_conditional and any(kw in cell_norm for kw in _REQUIRED_KW):
                    rule.obrigatorio = True
                elif any(kw in cell_norm for kw in _OPTIONAL_KW):
                    rule.obrigatorio = False

                # Skip internal codes (e.g. "ps_product_name|1|0") and opaque hashes
                is_code = _INTERNAL_CODE_RE.match(cell_str.lower().strip())
                is_hash = len(cell_str) <= 40 and " " not in cell_str and cell_str.isalnum()
                if len(cell_str) > 15 and cell_norm not in _REQUIRED_KW | _OPTIONAL_KW and not is_code and not is_hash:
                    if not rule.regra:
                        rule.regra = cell_str[:500]
                    vals = _extract_accepted_values(cell_str)
                    rule.valores_aceitos.extend(vals)

            rule.valores_aceitos = _dedup(rule.valores_aceitos)
            rules[col_name] = rule

        return rules

    # ── Definition sheet ──────────────────────────────────────────────────────

    def _parse_definition_sheet(
        self, ws, existing: dict[str, ColumnRule]
    ) -> dict[str, ColumnRule]:
        """
        Parseia aba de definição estilo Amazon/Walmart com tabela nome→regra.
        """
        rules: dict[str, ColumnRule] = {}
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return rules

        # Encontra linha de header da tabela
        hdr_idx = _find_table_header_row(rows)
        if hdr_idx is None:
            return rules

        col_hdrs = [_norm(str(v)) if v else "" for v in rows[hdr_idx]]

        name_i = _col_idx(col_hdrs, ["nome do campo", "attribute name", "campo", "field name", "field"])
        def_i  = _col_idx(col_hdrs, ["valores aceitos", "definicao", "definition", "definitions", "descricao"])
        ex_i   = _col_idx(col_hdrs, ["exemplo", "example values", "example"])
        req_i  = _col_idx(col_hdrs, ["required", "obrigatorio", "requirement level"])

        if name_i is None:
            return rules

        for row in rows[hdr_idx + 1:]:
            if not row or not row[name_i]:
                continue
            raw_name = str(row[name_i]).strip()
            if not raw_name or raw_name.startswith("("):
                continue

            # Tenta casar com coluna existente (nome técnico vs legível)
            matched = _match_col(raw_name, existing) or _match_col(raw_name, rules)
            col_name = matched or raw_name

            rule = existing.get(col_name) or rules.get(col_name) or ColumnRule(nome_coluna=col_name)

            if def_i is not None and def_i < len(row) and row[def_i]:
                rule.regra = str(row[def_i]).strip()[:500]
                rule.valores_aceitos.extend(_extract_accepted_values(rule.regra))

            if ex_i is not None and ex_i < len(row) and row[ex_i]:
                if not rule.exemplo:
                    rule.exemplo = str(row[ex_i]).strip()[:200]

            if req_i is not None and req_i < len(row) and row[req_i]:
                req_val = _norm(str(row[req_i]))
                if any(kw in req_val for kw in _REQUIRED_KW):
                    rule.obrigatorio = True

            rule.valores_aceitos = _dedup(rule.valores_aceitos)
            rules[col_name] = rule

        return rules

    # ── Example extraction ────────────────────────────────────────────────────

    def _extract_examples(self, ws, rules: dict[str, ColumnRule]) -> None:
        """
        Extrai valores de exemplo da aba e popula o campo ColumnRule.exemplo.
        Modifica rules in-place.
        """
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return

        # Header row: linha com mais valores nos primeiros 5 rows
        hdr_idx = 0
        best = 0
        for i, row in enumerate(rows[:5]):
            if not row:
                continue
            cnt = sum(1 for v in row if v and str(v).strip())
            if cnt > best:
                best, hdr_idx = cnt, i

        headers = [str(v).strip() if v else "" for v in rows[hdr_idx]]

        # Coleta valores das primeiras 5 linhas de exemplo
        for row in rows[hdr_idx + 1:hdr_idx + 6]:
            if not row:
                continue
            for col_idx, val in enumerate(row):
                if col_idx >= len(headers) or not headers[col_idx]:
                    continue
                if not val or not str(val).strip():
                    continue
                col_name = headers[col_idx]
                matched = _match_col(col_name, rules)
                target = matched or col_name
                rule = rules.get(target) or ColumnRule(nome_coluna=target)
                if not rule.exemplo:
                    rule.exemplo = str(val).strip()[:200]
                rules[target] = rule


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _norm(text: str) -> str:
    """Normaliza para lowercase sem acentos."""
    result = unicodedata.normalize("NFD", str(text))
    result = "".join(c for c in result if unicodedata.category(c) != "Mn")
    return result.strip().lower()


def _dedup(lst: list) -> list:
    return list(dict.fromkeys(lst))


def _merge_rule(base: ColumnRule, extra: ColumnRule) -> None:
    """Preenche campos vazios em base com dados de extra."""
    if not base.regra and extra.regra:
        base.regra = extra.regra
    if not base.valores_aceitos and extra.valores_aceitos:
        base.valores_aceitos = extra.valores_aceitos
    if not base.obrigatorio and extra.obrigatorio:
        base.obrigatorio = True
    if not base.exemplo and extra.exemplo:
        base.exemplo = extra.exemplo


def _find_table_header_row(rows: list) -> Optional[int]:
    """Encontra linha de cabeçalho de tabela de definições."""
    _KEYS = {
        "nome do campo", "attribute name", "campo", "field name",
        "definicao", "definition", "definitions",
        "nome como consta", "attribute",
    }
    for i, row in enumerate(rows[:12]):
        if not row:
            continue
        cells = {_norm(str(v)) for v in row if v}
        if cells & _KEYS:
            return i
    return None


def _col_idx(headers: list[str], keywords: list[str]) -> Optional[int]:
    """Retorna índice da primeira coluna que contenha algum keyword."""
    for kw in keywords:
        for i, h in enumerate(headers):
            if kw in h:
                return i
    return None


def _match_col(name: str, rules: dict) -> Optional[str]:
    """Casa nome técnico com nome legível via normalização e substring."""
    name_n = _norm(name)
    for col in rules:
        if _norm(col) == name_n:
            return col
        if name_n in _norm(col) or _norm(col) in name_n:
            return col
    return None


def _extract_accepted_values(text: str) -> list[str]:
    """
    Extrai lista de valores aceitos de texto de instrução.
    Detecta padrões como "Novo / Usado / Recondicionado" e "Valores: A, B, C".
    """
    if not text or len(text) < 4:
        return []

    # Padrão barra: "A / B / C" com itens curtos (< 40 chars cada)
    parts = [p.strip() for p in re.split(r"\s*/\s*", text)]
    if 2 <= len(parts) <= 15 and all(0 < len(p) <= 40 for p in parts):
        clean = [p.split("\n")[0].rstrip(".;,").strip() for p in parts]
        clean = [p for p in clean if p and not p.startswith("(")]
        if len(clean) >= 2:
            return clean

    # Padrão "Valores: A, B, C"
    m = re.search(
        r"(?:valores?|values?|opções?|options?)[:\s]+([^\n]{5,200})",
        text, re.IGNORECASE,
    )
    if m:
        chunk = m.group(1)
        items = [p.strip().rstrip(".;") for p in re.split(r"[,/]", chunk)]
        items = [it for it in items if 0 < len(it) <= 50]
        if len(items) >= 2:
            return items

    return []


def _load_wb(template_bytes: bytes):
    """Carrega workbook com fallback de sanitização."""
    import tempfile
    import warnings as _w

    import openpyxl

    _w.filterwarnings("ignore")
    try:
        from core.xlsx_openpyxl_compat import sanitize_xlsx_for_openpyxl
    except ImportError:
        sanitize_xlsx_for_openpyxl = lambda p: None  # noqa: E731

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(template_bytes)
        tmp_path = tmp.name

    san_path = sanitize_xlsx_for_openpyxl(tmp_path)
    # read_only=True breaks multi-column iteration on some xlsx files (e.g. Shopee)
    return openpyxl.load_workbook(
        san_path or tmp_path, data_only=True, keep_vba=True
    )
